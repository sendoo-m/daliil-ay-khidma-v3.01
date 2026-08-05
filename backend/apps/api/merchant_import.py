"""
Merchant — Bulk Product Import / Export
=======================================
تاجر عنده ٢٠٠ منتج لن يكتبها واحدًا واحدًا. هذه الوحدة تجعل الفارق بين
تبنّي التطبيق وهجره.

التصميم: **تصدير ← تعديل ← استيراد**.

لا نطلب من التاجر بناء ملف من الصفر. ينزّل منتجاته الحالية (أو قالبًا
بصف مثال لو لم يبدأ بعد)، يعدّل في إكسل الذي يعرفه، ويرفعه. عمود
"المعرّف" هو ما يجعل الرحلة دائرية: صف بمعرّف = تحديث، صف بلا معرّف
= إضافة.

والاستيراد على مرحلتين — فحص ثم حفظ. ملف من ٢٠٠ صف فيه ثلاثة أخطاء
لا يجوز أن يدخل نصفه ويترك النصف الآخر بلا إشعار.
"""

from __future__ import annotations

import csv
import io
from decimal import Decimal, InvalidOperation

from django.db import transaction

from apps.products.models import Product

#: حد أعلى للصفوف. ملف أكبر من هذا غالبًا خطأ لا نية.
MAX_ROWS = 1000

#: العناوين بالعربية — التاجر يفتح الملف ويفهمه بلا دليل.
#: المفتاح اسم الحقل في الموديل، والقيمة ما يراه في إكسل.
COLUMNS: dict[str, str] = {
    'id': 'المعرّف',
    'name_ar': 'الاسم',
    'description_ar': 'الوصف',
    'product_type': 'النوع',
    'price': 'السعر',
    'old_price': 'السعر قبل الخصم',
    'stock_quantity': 'الكمية',
    'is_available': 'متاح',
    'has_delivery': 'فيه توصيل',
    'delivery_cost': 'سعر التوصيل',
    'delivery_time_ar': 'وقت التوصيل',
}

HEADERS = list(COLUMNS.values())
HEADER_TO_FIELD = {label: field for field, label in COLUMNS.items()}

TYPE_TO_AR = {'product': 'منتج', 'service': 'خدمة'}
AR_TO_TYPE = {'منتج': 'product', 'خدمة': 'service',
              'product': 'product', 'service': 'service'}

TRUE_WORDS = {'نعم', 'ايوه', 'أيوه', 'صح', 'true', '1', 'yes', 'متاح'}
FALSE_WORDS = {'لا', 'لأ', 'غلط', 'false', '0', 'no', 'مش متاح', ''}


class ImportError_(Exception):
    """خطأ يمنع قراءة الملف أصلًا — لا يخص صفًا بعينه."""


# ═══════════════════════════════════════════════════════
#  التصدير
# ═══════════════════════════════════════════════════════

def _example_row() -> list:
    return [
        '',                     # المعرّف — فاضي يعني منتج جديد
        'مكرونة بشاميل',
        'مكرونة بالبشاميل والجبن، تكفي شخصين',
        'منتج',
        '75',
        '90',
        '20',
        'نعم',
        'نعم',
        '15',
        'من ٣٠ لـ٤٥ دقيقة',
    ]


def build_workbook(products) -> bytes:
    """ملف إكسل بمنتجات التاجر، أو قالب بصف مثال لو لا منتجات له."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover
        raise ImportError_('مكتبة openpyxl غير مثبّتة على الخادم.') from exc

    wb = Workbook()
    ws = wb.active
    ws.title = 'المنتجات'
    ws.sheet_view.rightToLeft = True  # ورقة عربية تُقرأ من اليمين

    ws.append(HEADERS)
    header_fill = PatternFill('solid', fgColor='14332B')
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.freeze_panes = 'A2'

    rows = list(products)
    if rows:
        for p in rows:
            ws.append([
                p.id,
                p.name_ar,
                p.description_ar,
                TYPE_TO_AR.get(p.product_type, 'منتج'),
                _decimal_str(p.price),
                _decimal_str(p.old_price),
                p.stock_quantity or 0,
                'نعم' if p.is_available else 'لا',
                'نعم' if p.has_delivery else 'لا',
                _decimal_str(p.delivery_cost),
                p.delivery_time_ar or '',
            ])
    else:
        ws.append(_example_row())

    widths = [10, 26, 40, 10, 10, 16, 10, 10, 12, 14, 20]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # ورقة تعليمات منفصلة: التاجر يقرأها مرة ولا تزحم بياناته.
    guide = wb.create_sheet('اقرأ ده الأول')
    guide.sheet_view.rightToLeft = True
    for line in [
        ['إزاي تستعمل الملف ده'],
        [''],
        ['١. عمود "المعرّف" متغيّرهوش.'],
        ['   صف فيه معرّف = تعديل منتج موجود.'],
        ['   صف المعرّف فيه فاضي = منتج جديد هيتضاف.'],
        [''],
        ['٢. الأعمدة المطلوبة: الاسم · الوصف · السعر'],
        [''],
        ['٣. "النوع" اكتب فيه: منتج أو خدمة'],
        [''],
        ['٤. أعمدة نعم/لا: اكتب نعم أو لا'],
        [''],
        ['٥. "السعر قبل الخصم" لازم يكون أعلى من "السعر".'],
        ['   سيبه فاضي لو مفيش خصم.'],
        [''],
        ['٦. الصور مش في الملف ده — تضيفها من التطبيق'],
        ['   بعد ما ترفع المنتجات.'],
        [''],
        [f'٧. أقصى عدد صفوف: {MAX_ROWS}'],
    ]:
        guide.append(line)
    guide.column_dimensions['A'].width = 60
    guide['A1'].font = Font(bold=True, size=13)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _decimal_str(value) -> str:
    """‏75.00 → 75 — الكسر الصفري ضجيج في ملف يقرأه إنسان."""
    if value in (None, ''):
        return ''
    try:
        d = Decimal(str(value))
    except (InvalidOperation, TypeError):
        return str(value)
    return str(d.quantize(Decimal('1')) if d == d.to_integral() else d)


# ═══════════════════════════════════════════════════════
#  القراءة
# ═══════════════════════════════════════════════════════

def read_rows(uploaded_file) -> list[dict]:
    """يقرأ xlsx أو csv ويعيد صفوفًا بمفاتيح حقول الموديل."""
    name = (getattr(uploaded_file, 'name', '') or '').lower()

    if name.endswith(('.xlsx', '.xlsm')):
        table = _read_xlsx(uploaded_file)
    elif name.endswith('.csv'):
        table = _read_csv(uploaded_file)
    else:
        raise ImportError_(
            'نوع الملف مش مدعوم. ارفع ملف Excel بامتداد .xlsx أو .csv'
        )

    if not table:
        raise ImportError_('الملف فاضي.')

    header = [str(c or '').strip() for c in table[0]]
    missing = [
        COLUMNS[f] for f in ('name_ar', 'description_ar', 'price')
        if COLUMNS[f] not in header
    ]
    if missing:
        raise ImportError_(
            'الملف ناقصه أعمدة: ' + ' · '.join(missing) +
            '. نزّل الملف من التطبيق واشتغل عليه.'
        )

    body = table[1:]
    if len(body) > MAX_ROWS:
        raise ImportError_(
            f'الملف فيه {len(body)} صف. الحد الأقصى {MAX_ROWS} صف '
            'في المرة الواحدة.'
        )

    rows = []
    for offset, raw in enumerate(body):
        # رقم الصف كما يراه التاجر في إكسل: العنوان صف ١.
        record = {'__row__': offset + 2}
        for index, label in enumerate(header):
            field = HEADER_TO_FIELD.get(label)
            if field is None:
                continue
            value = raw[index] if index < len(raw) else None
            record[field] = '' if value is None else str(value).strip()

        # نتجاهل الصفوف الفاضية تمامًا — إكسل يضيفها كثيرًا بلا قصد.
        if any(record.get(f) for f in ('name_ar', 'description_ar', 'price')):
            rows.append(record)

    if not rows:
        raise ImportError_('مفيش صفوف فيها بيانات.')
    return rows


def _read_xlsx(uploaded_file) -> list[list]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise ImportError_('مكتبة openpyxl غير مثبّتة على الخادم.') from exc

    try:
        wb = load_workbook(uploaded_file, data_only=True, read_only=True)
    except Exception as exc:
        raise ImportError_('تعذّر فتح الملف. اتأكد إنه Excel سليم.') from exc

    # ورقة "المنتجات" لو موجودة، وإلا الأولى — التاجر قد يعيد ترتيب الأوراق.
    ws = wb['المنتجات'] if 'المنتجات' in wb.sheetnames else wb.worksheets[0]
    return [list(r) for r in ws.iter_rows(values_only=True)]


def _read_csv(uploaded_file) -> list[list]:
    raw = uploaded_file.read()
    for encoding in ('utf-8-sig', 'utf-8', 'cp1256'):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ImportError_('ترميز الملف مش مفهوم. احفظه بترميز UTF-8.')
    return [row for row in csv.reader(io.StringIO(text))]


# ═══════════════════════════════════════════════════════
#  التحقق
# ═══════════════════════════════════════════════════════

def _to_decimal(value: str) -> Decimal | None:
    text = (value or '').replace(',', '').replace('٫', '.').strip()
    # الأرقام العربية الهندية — التاجر قد يكتبها من لوحة مفاتيح عربية.
    text = text.translate(str.maketrans('٠١٢٣٤٥٦٧٨٩', '0123456789'))
    if not text:
        return None
    try:
        return Decimal(text)
    except InvalidOperation:
        raise ValueError('مش رقم صحيح')


def _to_bool(value: str, default: bool = False) -> bool:
    text = (value or '').strip().lower()
    if text in TRUE_WORDS:
        return True
    if text in FALSE_WORDS:
        return default
    raise ValueError('اكتب نعم أو لا')


def validate_rows(rows: list[dict], owned_product_ids: set[int]) -> dict:
    """
    يفحص كل صف ويعيد تقريرًا. لا يكتب شيئًا.

    كل خطأ يحمل رقم الصف واسم العمود بالعربية، فيصلحه التاجر مباشرة
    في إكسل بدل أن يبحث عنه.
    """
    prepared: list[dict] = []
    errors: list[dict] = []
    seen_ids: set[int] = set()

    for row in rows:
        line = row['__row__']
        problems: list[dict] = []
        data: dict = {}

        # ── المعرّف ──
        raw_id = (row.get('id') or '').strip()
        product_id = None
        if raw_id:
            try:
                product_id = int(float(raw_id))
            except ValueError:
                problems.append({'field': COLUMNS['id'],
                                 'message': 'المعرّف لازم يكون رقم'})
            else:
                if product_id not in owned_product_ids:
                    problems.append({
                        'field': COLUMNS['id'],
                        'message': 'المنتج ده مش من منتجاتك أو مش موجود',
                    })
                elif product_id in seen_ids:
                    problems.append({'field': COLUMNS['id'],
                                     'message': 'المعرّف مكرر في الملف'})
                else:
                    seen_ids.add(product_id)

        # ── المطلوبة ──
        name = (row.get('name_ar') or '').strip()
        if len(name) < 2:
            problems.append({'field': COLUMNS['name_ar'],
                             'message': 'الاسم مطلوب'})
        data['name_ar'] = name
        data['name_en'] = name

        description = (row.get('description_ar') or '').strip()
        if not description:
            problems.append({'field': COLUMNS['description_ar'],
                             'message': 'الوصف مطلوب'})
        data['description_ar'] = description
        data['description_en'] = description

        # ── السعر ──
        price = old_price = None
        try:
            price = _to_decimal(row.get('price', ''))
            if price is None:
                problems.append({'field': COLUMNS['price'],
                                 'message': 'السعر مطلوب'})
            elif price <= 0:
                problems.append({'field': COLUMNS['price'],
                                 'message': 'السعر لازم يكون أكبر من صفر'})
        except ValueError as exc:
            problems.append({'field': COLUMNS['price'], 'message': str(exc)})

        try:
            old_price = _to_decimal(row.get('old_price', ''))
            if old_price is not None and price is not None and old_price <= price:
                problems.append({
                    'field': COLUMNS['old_price'],
                    'message': 'لازم يكون أعلى من السعر، أو سيبه فاضي',
                })
        except ValueError as exc:
            problems.append({'field': COLUMNS['old_price'],
                             'message': str(exc)})

        if price is not None:
            data['price'] = price
        data['old_price'] = old_price

        # ── النوع ──
        raw_type = (row.get('product_type') or '').strip()
        if raw_type and raw_type not in AR_TO_TYPE:
            problems.append({'field': COLUMNS['product_type'],
                             'message': 'اكتب: منتج أو خدمة'})
        data['product_type'] = AR_TO_TYPE.get(raw_type, 'product')

        # ── الاختيارية ──
        for field, label, default in (
            ('is_available', COLUMNS['is_available'], True),
            ('has_delivery', COLUMNS['has_delivery'], False),
        ):
            try:
                raw = (row.get(field) or '').strip()
                data[field] = default if not raw else _to_bool(raw, default)
            except ValueError as exc:
                problems.append({'field': label, 'message': str(exc)})

        try:
            stock = _to_decimal(row.get('stock_quantity', ''))
            data['stock_quantity'] = int(stock) if stock is not None else 0
        except ValueError:
            problems.append({'field': COLUMNS['stock_quantity'],
                             'message': 'الكمية لازم تكون رقم'})

        try:
            cost = _to_decimal(row.get('delivery_cost', ''))
            data['delivery_cost'] = cost if cost is not None else Decimal('0')
        except ValueError:
            problems.append({'field': COLUMNS['delivery_cost'],
                             'message': 'سعر التوصيل لازم يكون رقم'})

        data['delivery_time_ar'] = (row.get('delivery_time_ar') or '').strip()
        data['delivery_time_en'] = data['delivery_time_ar']

        if problems:
            errors.append({
                'row': line,
                'name': name or '—',
                'problems': problems,
            })
        else:
            prepared.append({'row': line, 'id': product_id, 'data': data})

    creates = sum(1 for p in prepared if p['id'] is None)
    return {
        'total_rows': len(rows),
        'valid': len(prepared),
        'will_create': creates,
        'will_update': len(prepared) - creates,
        'error_count': len(errors),
        'errors': errors[:50],   # تقرير أطول من ذلك لا يُقرأ
        'errors_truncated': len(errors) > 50,
        'prepared': prepared,
    }


# ═══════════════════════════════════════════════════════
#  الحفظ
# ═══════════════════════════════════════════════════════

@transaction.atomic
def commit_rows(prepared: list[dict], business) -> dict:
    """
    يكتب الصفوف السليمة. داخل transaction واحدة: إما تمام أو لا شيء.

    استيراد نصفه نجح يترك التاجر لا يعرف أين توقف، وإعادة رفع الملف
    تُنشئ نسخًا مكررة مما نجح.
    """
    created = updated = 0

    for item in prepared:
        data = item['data']
        if item['id'] is None:
            Product.objects.create(business=business, **data)
            created += 1
        else:
            Product.objects.filter(pk=item['id'], business=business).update(
                **data
            )
            updated += 1

    return {'created': created, 'updated': updated}
